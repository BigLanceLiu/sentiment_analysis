# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html


import pymysql
import openpyxl
from openpyxl import Workbook

class ExcelPipeline(object):
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(['comments'])

    def process_item(self, item, spider):

        line = item['comments'].replace('\n','')
        print(repr(line))
        self.ws.append([line])
        self.wb.save('darenwu.xlsx')
        return item
'''
class ExcelPipeline(object):
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(['movies'])

    def process_item(self, item, spider):

        line = item['movies'].replace('\n','')
        print(repr(line))
        self.ws.append([line])
        self.wb.save('.xlsx')
        return item
'''
'''
class ExcelPipeline(object):
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(['comments'])

    def process_item(self, item, spider):

        line = item['comments'].replace('\n','')
        print(repr(line))
        self.ws.append([line])
        self.wb.save('url.xlsx')
        return item
'''


'''
class MongoPipeline(object):

    collection = 'sansheng'    # 三生的数据表
    # collection = 'ershier'   # 二十二的数据表
    # collection = 'zhanlang'  # 战狼的数据表


    def __init__(self, mongo_uri, mongo_db):
        self.mongo_uri = mongo_uri
        self.mongo_db = mongo_db

    @classmethod
    def from_crawler(cls, crawler):
        return cls(
            mongo_uri = crawler.settings.get('MONGO_RUI'),
            mongo_db = crawler.settings.get('MONGO_DB')
        )

    def open_spider(self, spider):
        self.client = pymongo.MongoClient(self.mongo_uri)
        self.db = self.client[self.mongo_db]

    def close_spider(self, spider):
        self.client.close()

    def process_item(self, item, spider):
        table = self.db[self.collection]
        for com in item['comments']:
            data = dict()
            data['comment'] = com.strip().replace("\n", "")
            table.insert_one(data)
        return item
'''